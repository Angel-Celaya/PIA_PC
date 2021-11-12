from bs4 import BeautifulSoup
from pyhunter import PyHunter
from openpyxl import Workbook
from scapy.all import *
from lxml import html
import requests
import logging
import getpass
import socket
import sys
import os

logging.basicConfig(filename='app.log',level=logging.INFO) 
#------------------------------Funciones para el cifrado----------------------------#
def ccon_clave(message,key):
    SYMBOLS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz1234567890 !?.'
    translated = ''
    for symbol in message:
        if symbol in SYMBOLS:
            symbolIndex = SYMBOLS.find(symbol)
            translatedIndex = symbolIndex + key
            
            if translatedIndex >= len(SYMBOLS):
                translatedIndex = translatedIndex - len(SYMBOLS)
            elif translatedIndex < 0:
                translatedIndex = translatedIndex + len(SYMBOLS)

            translated = translated + SYMBOLS[translatedIndex]
        else:
            translated = translated + symbol
    logging.info("Se cifro un mensaje")
    print("El mensaje cifrado es: ",translated)

def dcon_clave(messagec,key):
    SYMBOLS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz1234567890 !?.'
    translated = ''
    for symbol in messagec:
        if symbol in SYMBOLS:
            symbolIndex = SYMBOLS.find(symbol)
            translatedIndex = symbolIndex - key
            
            if translatedIndex >= len(SYMBOLS):
                translatedIndex = translatedIndex - len(SYMBOLS)
            elif translatedIndex < 0:
                translatedIndex = translatedIndex + len(SYMBOLS)

            translated = translated + SYMBOLS[translatedIndex]
        else:
            translated = translated + symbol
    logging.info("Se descifro un mensaje")
    print("El mensaje descifrado es: ",translated)
    
def crack(messagec):
    SYMBOLS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz1234567890 !?.'
    translated = ''
    for key in range(len(SYMBOLS)):
        translated = ''
        for symbol in messagec:
            if symbol in SYMBOLS:
                symbolIndex = SYMBOLS.find(symbol)
                translatedIndex = symbolIndex - key
                if translatedIndex < 0:
                    translatedIndex = translatedIndex + len(SYMBOLS)
                translated = translated + SYMBOLS[translatedIndex]
            else:
                translated = translated + symbol
        print('Key #%s: %s' % (key, translated))
    print()
    logging.info("Se descifro un mensaje mediante crackeo")
    crackeo(messagec)

MAYUSCULAS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZ'
LETRAS_Y_ESPACIOS = MAYUSCULAS + MAYUSCULAS.lower() + ' \t\n'

def loadDictionary():
    dictionaryFile = open('dictEsp.txt', encoding='utf-8')
    englishWords = {}
    for word in dictionaryFile.read().split('\n'):
        word = word.upper()
        englishWords[word] = None
    dictionaryFile.close()
    return englishWords

SPANISH_WORDS = loadDictionary()

def getSpanishCount(message):
    message = message.upper()
    message = removeNonLetters(message)
    possibleWords = message.split()
    if possibleWords == []:
        return 0.0

    matches = 0
    for word in possibleWords:
        if word in SPANISH_WORDS:
            matches += 1
    return float(matches) / len(possibleWords)


def removeNonLetters(message):
    lettersOnly = []
    for symbol in message:
        if symbol in LETRAS_Y_ESPACIOS:
            lettersOnly.append(symbol)
    return ''.join(lettersOnly)


def isSpanish(message, wordPercentage=60, letterPercentage=85):
    wordsMatch = getSpanishCount(message) * 100 >= wordPercentage
    numLetters = len(removeNonLetters(message))
    messageLettersPercentage = float(numLetters) / len(message) * 100
    lettersMatch = messageLettersPercentage >= letterPercentage
    return wordsMatch and lettersMatch

def crackeo(mensaje):
    SYMBOLS = 'ABCDEFGHIJKLMNOPQRSTUVWXYZabcdefghijklmnopqrstuvwxyz1234567890 !?.'
    for key in range(1, len(SYMBOLS)):
        trad = ''
        for abc in mensaje:
            if abc in SYMBOLS:
                abcIndex = SYMBOLS.find(abc)
                tradIndex = abcIndex - key
                if tradIndex < 0:
                    tradIndex = tradIndex + len(SYMBOLS)
                trad = trad + SYMBOLS[tradIndex]
            else:
                trad = trad + abc

        if isSpanish(trad):
            print("Key:", str(key)+". Mensaje encontrado: "+trad[:100])
            print("Válido para idioma español")

#-------------------------------Funcion para el escaneo de puertos---------------------#
def checkPortsSocket(ip,portlist):
    print("IP",ip,type(ip))
    try:
        for port in portlist:
            sock=socket.socket(socket.AF_INET,socket.SOCK_STREAM)
            sock.settimeout(5)
            result=sock.connect_ex((ip,port))
            if result == 0:
                print("Puerto {}: \t Abierto".format(port))
            else:
                print("Puerto {}: \t Cerrado".format(port))
            sock.close()
        logging.info("Se utilizo el escaneo de puertos ")
    except socket.error as error:
        logging.error("Ha ocurrido un error "+str(e))
        print(str(error))
        print("Error de conexion")
        sys.exit()

#-------------------------------Funciones para el Web scraping---------------------#
def opcionwebscraping(url):
    #print(url)
    option=input("Opncion: ")
    if option=="img":
        scrapingBeautifulSoup(url)
    elif option=="pdf":
        scrapingPDF(url)
    elif option=="link":
        scrapingLinks(url)
    else:
        print("Opcion no valida\n")
        opcionwebscraping(url)

def scrapingBeautifulSoup(url):
    try:
        print("Obteniendo imagenes con BeautifulSoup "+ url)
        response = requests.get(url)
        bs = BeautifulSoup(response.text, 'lxml')
        #create directory for save images
        os.system("mkdir images")
        for tagImage in bs.find_all("img"): 
            #print(tagImage['src'])
            if tagImage['src'].startswith("http") == False:
                download = url + tagImage['src']
            else:
                download = tagImage['src']
            print(download)
            # download images in img directory
            r = requests.get(download)
            f = open('images/%s' % download.split('/')[-1], 'wb')
            f.write(r.content)
            f.close()
        logging.info("Se descargaron imagenes con webscraping")    
    except Exception as e:
        print(e)
        print("Error conexion " + url)
        logging.error("Ha ocurrido un error "+str(e))
        pass

def scrapingPDF(url):
    print("\nObteniendo pdfs de la url:"+ url)
    try:
        response = requests.get(url)  
        parsed_body = html.fromstring(response.text)
        # expresion regular para obtener pdf
        pdfs = parsed_body.xpath('//a[@href[contains(., ".pdf")]]/@href')
        #create directory for save pdfs
        if len(pdfs) >0:
            os.system("mkdir pdfs")
        print ('Encontrados %s pdf' % len(pdfs))    
        for pdf in pdfs:
            if pdf.startswith("http") == False:
                download = url + pdf
            else:
                download = pdf
            print(download)
                    
            # descarga pdfs
            r = requests.get(download)
            f = open('pdfs/%s' % download.split('/')[-1], 'wb')
            f.write(r.content)
            f.close()
        logging.info("Se descargaron PDFs con webscraping")    
    except Exception as e:
        print(e)
        print("Error conexion con " + url)
        logging.error("Ha ocurrido un error "+str(e))
        pass

def scrapingLinks(url):
    print("\nObteniendo links de la url:"+ url)
        
    try:
        response = requests.get(url)  
        parsed_body = html.fromstring(response.text)
        # expresion regular para obtener links
        links = parsed_body.xpath('//a/@href')
        print('links %s encontrados' % len(links))
        for link in links:
            print(link)
        logging.info("Se mostraron enlaces con webscraping")           
    except Exception as e:
        print(e)
        print("Error conexion con " + url)
        logging.error("Ha ocurrido un error "+str(e))
        pass
        
#-------------------------------Funciones para el scapy---------------------#
def ip_scan(ip):
    range_ip=ARP(pdst=ip)
    broadcast=Ether(dst="ff:ff:ff:ff:ff:ff")
    final_packet=broadcast/range_ip
    res=srp(final_packet, timeout=2, verbose=False)[0]
    for n in res:
        print("[+]host: {}     MAC: {}".format(n[1].psrc,n[1].hwsrc))
    logging.info("Se escaneo todos los dispositivos conectados a la misma red")

#------------------------------Funciones para el hunter-----------------------#
def Busqueda(orga):
        # Cantidad de resultados esperados de la búsqueda
        # El límite MENSUAL de Hunter es 50, cuidado!
    resultado = hunter.domain_search(company=organizacion, limit=1, emails_type="personal")
    return resultado

def GuardarInformacion(datosEncontrados, organizacion):
    # libro = openxyl.Workbook()
    libro = Workbook()
    pagina = libro.active
    pagina1 = libro.create_sheet(organizacion)
    if len(libro.sheetnames) > 1:
        libro.active = 1
        hoja = libro.active
    else:
        hoja = libro.active
    print("La página activa es:", hoja.title)
    datosImportantes = ("domain", "organization", "country", "emails")
    datosImportantes2 = (
        "value", "type", "sources", "first_name", "last_name", "phone_number"
    )
    datosImportantes3 = ("domain", "uri", "extracted_on", "last_seen_on")
    data1 = list()
    data2 = list()
    data3 = list()
    hoja["A1"] = "domain"
    for i in range(1, 4):
        hoja[f"A{1+i}"] = datosImportantes[i]
    for val in range(len(datosImportantes2)):
        hoja[f"A{4+val}"] = datosImportantes2[val]
    hoja["A6"] = "first_name"
    hoja["A7"] = "last_name"
    hoja["A8"] = "phone_number"
    for val2 in range(len(datosImportantes3)):
        hoja[f"A{8+val2}"] = datosImportantes3[val2]
    for y in datosImportantes:
        x = datosEncontrados[y]
        data1.append(x)
        value = type(x).__name__
        if value == "list":
            newdata = datosEncontrados[y]
            for data in datosImportantes2:
                for elem in newdata:
                    i = elem[data]
                    value2 = type(i).__name__
                    data2.append(i)
                    if value2 == "list":
                        lista = elem[data]
                        for elem2 in datosImportantes3:
                            for f in lista:
                                z = f[elem2]
                                data3.append(z)
    data1.pop()
    a = 1
    for i in data1:
        hoja[f"B{a}"] = i
        a += 1
    data2.pop(2)
    a = 4
    for i in data2:
        hoja[f"B{a}"] = i
        a += 1
    a = 8
    for i in data3:
        hoja[f"B{a}"] = i
        a += 1
    libro.save("Hunter" + organizacion + ".xlsx")
